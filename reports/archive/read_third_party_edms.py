"""Probe the third-party eDM xlsx/xls files in OneDrive to find their structure."""
from pathlib import Path
import sys

import openpyxl
import xlrd

ONEDRIVE = Path("/mnt/c/Users/User/Downloads/OneDrive_2026-04-29/5. Media Metrics")


def probe_xlsx(p: Path):
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"  [{sheet}]  dims={ws.calculate_dimension()}")
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append([str(c) if c is not None else "" for c in row])
                if i >= 8:
                    break
            for r in rows:
                print("    " + " | ".join(c[:25] for c in r[:8]))
        wb.close()
    except Exception as e:
        print(f"  xlsx err: {e}")


def probe_xls(p: Path):
    try:
        wb = xlrd.open_workbook(p, on_demand=True)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            print(f"  [{sheet_name}]  rows={ws.nrows} cols={ws.ncols}")
            for i in range(min(8, ws.nrows)):
                row = [str(ws.cell_value(i, j))[:25] for j in range(min(8, ws.ncols))]
                print("    " + " | ".join(row))
    except Exception as e:
        print(f"  xls err: {e}")


def main():
    for p in sorted(ONEDRIVE.rglob("*.xls*")):
        rel = p.relative_to(ONEDRIVE)
        print(f"\n=== {rel} ({p.stat().st_size:,} bytes) ===")
        if p.suffix.lower() == ".xlsx":
            probe_xlsx(p)
        elif p.suffix.lower() == ".xls":
            probe_xls(p)


if __name__ == "__main__":
    main()
