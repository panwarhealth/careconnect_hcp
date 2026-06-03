"""Parse all third-party eDM xlsx/xls files into a single metrics CSV.

    python reports/edm/extract.py
"""
import csv
import sys
from pathlib import Path

import openpyxl
import xlrd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from lib import io

ONEDRIVE = Path("/mnt/c/Users/User/Downloads/OneDrive_2026-04-29/5. Media Metrics")
OUT = io.out_dir("third_party_edms")


def num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).strip().rstrip("%")
        try:
            return float(s)
        except ValueError:
            return None


def parse_solus_xls(p: Path) -> dict:
    """Solus .xls files are key/value pairs in a 3-col layout."""
    wb = xlrd.open_workbook(p, on_demand=True)
    ws = wb.sheet_by_index(0)
    data = {"source_file": p.name}
    for r in range(ws.nrows):
        key = str(ws.cell_value(r, 0)).strip().rstrip(":").lower()
        val = ws.cell_value(r, 1)
        if not key or val == "":
            continue
        data[key] = val
    return data


def parse_tmr_xlsx(p: Path) -> list[dict]:
    """TMR/MedicalRepublic xlsx files have a header block then a table."""
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows):
        if row and "TitleID" in (str(c) for c in row if c is not None):
            header_row = i
            break
    if header_row is None:
        return []
    headers = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(rows[header_row])]
    out = []
    for r in rows[header_row + 1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        d = {h: r[i] if i < len(r) else None for i, h in enumerate(headers)}
        d["source_file"] = p.name
        out.append(d)
    return out


def parse_princeton(p: Path) -> list[dict]:
    """Princeton 'Monthly Sessions' xlsx has a content × month grid."""
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    brand = ""
    for r in rows[:5]:
        if r and r[0] and "Brand:" in str(r[0]):
            brand = str(r[0]).split(":", 1)[1].strip()
    header_row = None
    for i, r in enumerate(rows):
        if r and r[0] and "Content Sessions" in str(r[0]):
            header_row = i
            break
    if header_row is None:
        return []
    months = [str(c)[:10] if c else "" for c in rows[header_row][1:]]
    out = []
    for r in rows[header_row + 1:]:
        if not r or not r[0]:
            continue
        content = str(r[0]).strip()
        if content.lower() in ("fm months", ""):
            continue
        for m, val in zip(months, r[1:]):
            if val is None or val == "":
                continue
            out.append({"brand": brand, "content": content, "month": m, "sessions": val, "source_file": p.name})
    return out


def main():
    OUT.mkdir(parents=True, exist_ok=True)

    solus_rows = []
    for p in sorted(ONEDRIVE.rglob("*.xls")):
        try:
            d = parse_solus_xls(p)
            solus_rows.append(d)
        except Exception as e:
            print(f"err {p.name}: {e}")

    tmr_rows = []
    princeton_rows = []
    for p in sorted(ONEDRIVE.rglob("*.xlsx")):
        if "Princeton" in str(p) or "Monthly Sessions" in p.name:
            princeton_rows.extend(parse_princeton(p))
        elif "TMR" in p.name or "Solus EDM report" in p.name:
            tmr_rows.extend(parse_tmr_xlsx(p))

    if solus_rows:
        # Tidy headline metrics per campaign
        tidy_rows = []
        wanted = ["title", "subject line", "delivery date/time", "total recipients", "successful deliveries", "total opens", "recipients who opened", "total clicks", "recipients who clicked", "source_file"]
        for d in solus_rows:
            tidy_rows.append({k: d.get(k, "") for k in wanted})
        with (OUT / "solus_campaigns.csv").open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=wanted)
            w.writeheader()
            w.writerows(tidy_rows)
        print(f"wrote solus_campaigns.csv  ({len(tidy_rows)} campaigns)")

    if tmr_rows:
        keys = sorted({k for d in tmr_rows for k in d.keys()})
        with (OUT / "tmr_campaigns.csv").open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(tmr_rows)
        print(f"wrote tmr_campaigns.csv  ({len(tmr_rows)} rows)")

    if princeton_rows:
        keys = ["brand", "content", "month", "sessions", "source_file"]
        with (OUT / "princeton_desk_sets.csv").open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(princeton_rows)
        print(f"wrote princeton_desk_sets.csv  ({len(princeton_rows)} rows)")


if __name__ == "__main__":
    main()
